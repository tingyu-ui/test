#!/usr/bin/env python
# -*- coding:utf-8 -*-

from openpyxl import load_workbook
wb = load_workbook(filename = 'empty_boo1.xlsx')
sheet_ranges = wb['range names']
print(sheet_range['D18'].value)
for i in range(1,31):
    print(sheet_range.cell(column=1,row=i).value)

